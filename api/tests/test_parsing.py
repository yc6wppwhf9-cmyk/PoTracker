"""Tests for the RM sheet parser.

The parser is the riskiest code in the project — heuristic header detection,
scored alias matching, and value-based fallback column inference — so it is the
piece most worth pinning down.
"""
from __future__ import annotations

import io

import openpyxl
import pytest

from app.parsing import parse_rm_sheet


def _workbook(rows: list[list], title: str = "PLANT-A") -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


HEADER = [
    "COMPONENT ICODE",
    "COMPONENT ITEM NAME",
    "Final Pending (To Purchase)",
    "UOM",
    "Lot No",
    "Department",
    "COLOR",
    "PO Number",
]


def test_parses_header_driven_rows():
    data = _workbook([
        HEADER,
        ["INV20605", "Zipper 8in", 120, "pcs", "LOT-1", "Trims", "Black", "PO-77"],
        ["ADJ12", "Adjuster", 40, "pcs", "LOT-2", "Trims", "Silver", None],
    ])
    parsed = parse_rm_sheet(data)

    assert len(parsed.rows) == 2
    first = parsed.rows[0]
    assert first.raw_code == "INV20605"
    assert first.item_name == "Zipper 8in"
    assert first.required_qty == 120
    assert first.location == "PLANT-A"  # tab name


def test_populates_unit_po_and_color():
    """These three fields drive UOM conversion and auto-PO extraction; if the
    parser stops filling them, those features silently become no-ops."""
    data = _workbook([
        HEADER,
        ["INV20605", "Zipper 8in", 120, "mtr", "LOT-1", "Trims", "Black", "PO-77"],
    ])
    row = parse_rm_sheet(data).rows[0]

    assert row.raw_unit == "mtr"
    assert row.po == "PO-77"
    assert row.color == "Black"
    assert row.lot == "LOT-1"
    assert row.department == "Trims"


def test_skips_non_positive_quantities():
    """Pending = required - stock. Zero means exactly covered; negative means
    surplus. Neither is a purchase, and a negative stored as demand would net
    off other items' genuine requirement in any category total."""
    data = _workbook([
        HEADER,
        ["INV1001", "Covered by stock", 0, "pcs", None, None, None, None],
        ["INV1002", "No qty", None, "pcs", None, None, None, None],
        ["INV1003", "To buy", 5, "pcs", None, None, None, None],
        ["INV1004", "Surplus stock", -16592, "pcs", None, None, None, None],
    ])
    parsed = parse_rm_sheet(data)

    assert [r.raw_code for r in parsed.rows] == ["INV1003"]
    assert parsed.skipped_rows == 3
    assert all(r.required_qty > 0 for r in parsed.rows)


def test_surplus_row_from_the_real_sheet_is_not_demand():
    """INV26643: need 10,240, hold 26,832 in stock, pending -16,592.
    Nothing to buy — the line must not enter the requirement set, and must not
    net against the genuine demand on the line beside it."""
    data = _workbook([
        ["COMPONENT ICODE", "COMPONENT ITEM NAME", "TOTAL_COMPONENT_QUANTITY",
         "SOH", "Pending"],
        ["INV26643", "WHL Atlas BL FSN", 10240, 26832, -16592],
        ["INV1371", "Wheel Cover", 7000, 0, 7000],
    ])
    parsed = parse_rm_sheet(data)

    assert [r.raw_code for r in parsed.rows] == ["INV1371"]
    assert sum(r.required_qty for r in parsed.rows) == 7000


def test_fully_stocked_sheet_reports_why_rather_than_a_format_error():
    """A sheet with nothing left to buy is a valid sheet, not a malformed one."""
    data = _workbook([
        ["COMPONENT ICODE", "COMPONENT ITEM NAME", "Pending"],
        ["INV26643", "WHL Atlas BL FSN", -16592],
    ])
    with pytest.raises(ValueError, match="Nothing to purchase"):
        parse_rm_sheet(data)


def test_header_not_on_first_row():
    data = _workbook([
        ["HSCVPL — Raw Material Requirement"],
        [],
        HEADER,
        ["INV20605", "Zipper 8in", 120, "pcs", None, None, None, None],
    ])
    assert len(parse_rm_sheet(data).rows) == 1


def test_falls_back_to_total_quantity_column():
    data = _workbook([
        ["COMPONENT ICODE", "COMPONENT ITEM NAME", "TOTAL_COMPONENT_QUANTITY"],
        ["INV20605", "Zipper 8in", 300],
    ])
    row = parse_rm_sheet(data).rows[0]
    assert row.required_qty == 300


def test_detects_item_code_column_by_value_when_header_is_junk():
    data = _workbook([
        ["`", "COMPONENT ITEM NAME", "Final Pending (To Purchase)"],
        ["INV20605", "Zipper 8in", 10],
        ["ADJ12", "Adjuster", 20],
        ["ZIP8", "Zip", 30],
    ])
    parsed = parse_rm_sheet(data)
    assert [r.raw_code for r in parsed.rows] == ["INV20605", "ADJ12", "ZIP8"]


def test_multi_tab_workbook_merges_rows_and_tags_location():
    wb = openpyxl.Workbook()
    a = wb.active
    a.title = "PLANT-A"
    a.append(HEADER)
    a.append(["INV1", "A", 1, "pcs", None, None, None, None])
    b = wb.create_sheet("PLANT-B")
    b.append(HEADER)
    b.append(["INV2", "B", 2, "pcs", None, None, None, None])
    buf = io.BytesIO()
    wb.save(buf)

    parsed = parse_rm_sheet(buf.getvalue())
    assert {r.location for r in parsed.rows} == {"PLANT-A", "PLANT-B"}


def test_prefers_component_columns_over_finished_goods_columns():
    """Shape of the real 'Total DGPNT' tab: a finished-good code and order qty
    sit alongside the component code and component qty. The requirement is the
    *component*, so picking FG_ITEM_CODE / ORDERED_QUANTITY would be wrong."""
    data = _workbook([
        [
            "FG_ITEM_CODE",
            "FG_ITEM_NAME",
            "ORDERED_QUANTITY",
            "COMPONENT ICODE",
            "COMPONENT ITEM NAME",
            "TOTAL_COMPONENT_QUANTITY",
        ],
        ["INV29086", "DIY Universal 006", 408, "INV21780", "FB Bolt 170 WHT", 25.296],
    ])
    row = parse_rm_sheet(data).rows[0]

    assert row.raw_code == "INV21780"
    assert row.required_qty == pytest.approx(25.296)


def test_prefers_pending_over_total_when_both_present():
    """'Pending' is net of stock on hand; 'Total' is gross. Using the gross
    column when a pending column exists would over-order every line."""
    data = _workbook([
        [
            "COMPONENT ICODE",
            "TOTAL_COMPONENT_QUANTITY",
            "SOH_COMPONENT_QUANTITY",
            "PENDING_COMPONENT_TO_ISSUE",
        ],
        ["INV20605", 1000, 400, 600],
    ])
    row = parse_rm_sheet(data).rows[0]
    assert row.required_qty == 600


def test_raises_when_no_recognisable_columns():
    data = _workbook([["Alpha", "Beta"], ["x", "y"]])
    with pytest.raises(ValueError, match="No requirement rows found"):
        parse_rm_sheet(data)
