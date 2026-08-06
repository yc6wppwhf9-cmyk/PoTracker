"""Parse the GRN (GRC) register export.

The register is a flat sheet, one row per receipt line, with these columns:

    Stock Point Name, Supplier Name, GRC NO., GRC DATE, DOC DATE, DOC NO,
    PO DATE, PO NO., LOT NO., DEPARTMENT, BARCODE, ITEM_NAME, QTY,
    Per Item Cost (Landed Value), Landed Cost, Total Value of Invoice,
    Basic Amt., IGST Rate, IGST Amount, CGST Rate, CGST Amount, SGST Rate,
    SGST Amount, Other Charges, Freight, Dispatch, Remarks

Only the identifying columns and QTY are needed; the tax and value columns
are ignored.

Header aliases are matched loosely, as with the RM sheet parser, because these
exports vary between runs — trailing dots, case, and the odd renamed column are
normal and should not require a code change.

Memory matters here. An earlier version kept every cell of every row as a JSON
blob for diagnosis, which held a second copy of the whole spreadsheet in memory
and sent it to the database as well — enough to exhaust a small instance and
have the process killed mid-upload. The source file is the diagnosis; the rows
carry only what is used.
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import load_workbook

# Longest aliases are tried first so "po date" cannot be swallowed by "po".
HEADER_ALIASES: dict[str, list[str]] = {
    "grc_no": ["grc no", "grc number", "grn no", "grn number", "grc", "grn"],
    "grc_date": ["grc date", "grn date", "receipt date"],
    "po_number": ["po no", "po number", "purchase order no", "order no"],
    "po_date": ["po date", "purchase order date"],
    "item_code": ["barcode", "item barcode", "component icode", "icode", "item code"],
    "item_name": ["item_name", "item name", "component item name", "description"],
    "lot": ["lot no", "lot number", "lot"],
    "qty": ["qty", "quantity", "received qty", "grn qty", "accepted qty"],
    "stock_point": ["stock point name", "stock point", "site", "plant"],
    "supplier": ["supplier name", "supplier", "vendor name", "vendor"],
    "department": ["department", "dept"],
    "doc_no": ["doc no", "document no", "invoice no"],
    "doc_date": ["doc date", "document date", "invoice date"],
    "landed_cost": ["landed cost", "per item cost (landed value)", "landed value"],
    "remarks": ["remarks", "remark", "note"],
}

# A column whose name merely CONTAINS one of these is too weak to accept — they
# appear inside several headers ("Total Value of Invoice" contains "invoice").
_EXACT_ONLY = {"qty", "lot", "grc", "grn", "supplier", "vendor", "department"}


def _norm(value: Any) -> str:
    return re.sub(r"[^a-z0-9 ]+", " ", str(value or "").strip().lower()).strip()


def match_header(cell: Any) -> Optional[str]:
    """Map one header cell to a field name, or None."""
    n = _norm(cell)
    if not n:
        return None
    n = re.sub(r"\s+", " ", n)

    for field, aliases in HEADER_ALIASES.items():
        if n in [_norm(a) for a in aliases]:
            return field

    # Substring fallback, longest alias first, and never for the ambiguous ones.
    for field, aliases in sorted(
        HEADER_ALIASES.items(),
        key=lambda kv: max(len(a) for a in kv[1]),
        reverse=True,
    ):
        for a in aliases:
            if a in _EXACT_ONLY:
                continue
            if len(a) >= 4 and _norm(a) in n:
                return field
    return None


def _num(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    cleaned = re.sub(r"[,\s]", "", str(value))
    try:
        return float(cleaned)
    except ValueError:
        return None


def _date(value: Any) -> Optional[str]:
    """ISO date string, or None. Accepts what Excel and humans both produce."""
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d", "%d-%b-%Y", "%d %b %Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _clean_po_number(value: Any) -> Optional[str]:
    """Normalise a PO number for matching.

    Uppercased and stripped of internal spaces, because the same number is
    typed by hand in two systems and 'PO/HSM/00000910/26-27' must match
    'po/hsm/00000910/26-27 '. Nothing else is altered — the separators are
    part of the number.
    """
    if value in (None, ""):
        return None
    text = re.sub(r"\s+", "", str(value)).upper()
    return text or None


class GrnRow(dict):
    """A parsed receipt line. A dict so it can go straight to the database."""


def parse_grn_register(data: bytes) -> dict[str, Any]:
    """Parse the register.

    Returns the rows, plus counts explaining anything skipped — a silent skip
    is how a missing delivery becomes invisible.
    """
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    # Merged as we go. Collecting every row and deduplicating afterwards held
    # two copies of the register at once, which is what a large export cannot
    # afford on a small instance.
    merged: dict[tuple[str, str, str], GrnRow] = {}
    parsed_rows = 0
    merged_count = 0
    skipped_no_grc = 0
    skipped_no_qty = 0
    header_found = False
    headers_seen: list[str] = []

    for ws in wb.worksheets:
        mapping: dict[int, str] = {}
        for raw_row in ws.iter_rows(values_only=True):
            if not raw_row or all(c in (None, "") for c in raw_row):
                continue

            if not mapping:
                # The first row that maps enough columns is the header. GRC
                # exports often carry a title row or two above it.
                candidate = {
                    i: f
                    for i, c in enumerate(raw_row)
                    if (f := match_header(c)) is not None
                }
                if {"grc_no", "item_code", "qty"} <= set(candidate.values()) or (
                    len(candidate) >= 6 and "qty" in candidate.values()
                ):
                    mapping = candidate
                    header_found = True
                    headers_seen = [str(c) for c in raw_row if c not in (None, "")]
                continue

            record: dict[str, Any] = {}
            for i, cell in enumerate(raw_row):
                field = mapping.get(i)
                if field:
                    record[field] = cell

            grc_no = str(record.get("grc_no") or "").strip()
            if not grc_no:
                skipped_no_grc += 1
                continue

            qty = _num(record.get("qty"))
            # A zero or negative receipt is a correction, not goods arriving;
            # counting it would overstate what has landed.
            if qty is None or qty <= 0:
                skipped_no_qty += 1
                continue

            item_code = str(record.get("item_code") or "").strip().upper() or None
            lot = str(record.get("lot") or "").strip() or None

            parsed_rows += 1
            key = (grc_no, item_code or "", lot or "")
            existing = merged.get(key)
            if existing is not None:
                # The same receipt line split across two rows of the export is
                # one arrival, not two.
                existing["qty"] += qty
                merged_count += 1
                continue

            merged[key] = GrnRow(
                grc_no=grc_no,
                grc_date=_date(record.get("grc_date")),
                po_number=_clean_po_number(record.get("po_number")),
                po_date=_date(record.get("po_date")),
                item_code=item_code,
                item_name=str(record.get("item_name") or "").strip() or None,
                lot=lot,
                qty=qty,
                stock_point=str(record.get("stock_point") or "").strip() or None,
                supplier=str(record.get("supplier") or "").strip() or None,
                department=str(record.get("department") or "").strip() or None,
                doc_no=str(record.get("doc_no") or "").strip() or None,
                doc_date=_date(record.get("doc_date")),
                landed_cost=_num(record.get("landed_cost")),
                remarks=str(record.get("remarks") or "").strip() or None,
            )

    wb.close()

    return {
        "rows": list(merged.values()),
        "parsed": parsed_rows,
        "merged": merged_count,
        "skipped_no_grc": skipped_no_grc,
        "skipped_no_qty": skipped_no_qty,
        "header_found": header_found,
        "headers": headers_seen,
    }
