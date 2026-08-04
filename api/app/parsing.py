"""Deterministic parsing of an RM requirement workbook (.xlsx) into raw rows.

Handles multi-tab workbooks with varying layouts:
  * header not necessarily on row 1;
  * item-code column may have a junk header (e.g. a backtick) — detected by value;
  * required qty = the "Pending (to purchase)" column, falling back to the
    "Total component quantity" column when a tab has no pending column;
  * rows whose chosen quantity is 0/blank are skipped (already covered by stock).

No DB access and no LLM here — just structural extraction. Item-code matching
and fuzzy resolution happen later.
"""
from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from typing import Any, Optional

import openpyxl

# item codes look like INV20605, ADJ12, ZIP8 ...
CODE_RE = re.compile(r"^[A-Za-z]{2,5}\d{2,}[A-Za-z0-9-]*$")

# field -> candidate header aliases (normalised: lowercased, underscores→spaces,
# whitespace collapsed). Matched by exact (strong) or substring (weak); the
# longest matching alias wins, so specific beats generic.
ALIASES: dict[str, list[str]] = {
    "item_code": ["component icode", "icode", "component code", "`", "'", "code", "item code", "sku"],
    "item_name": ["component item name", "item name", "description", "particulars"],
    "pending_qty": [
        "pending component to issue",
        "pending to issue",
        "final pending",
        "to purchase",
        "pending",
    ],
    "total_qty": [
        "total component quantity",
        "total qty",
        "total quantity",
        "required qty",
    ],
    "soh": ["soh component quantity", "soh", "stock on hand", "stock"],
    "lot": ["lot no", "lot"],
    "department": ["department", "dept"],
    "channel": ["channel", "poc"],
    "remark": ["remark from tj", "remarks", "remark"],
    # Native unit of the quantity column; converted to the catalogue base unit
    # later via uom_conversion.
    "unit": ["unit of measure", "uom", "unit", "units"],
    # A PO number already present on the sheet (pre-placed orders).
    "po": ["po number", "po no", "po #", "purchase order"],
    "color": ["colour", "color"],
}


@dataclass
class ParsedRow:
    row_number: int
    raw_code: Optional[str]
    item_name: Optional[str]
    required_qty: Optional[float]
    raw_unit: Optional[str] = None
    lot: Optional[str] = None
    department: Optional[str] = None
    color: Optional[str] = None
    location: Optional[str] = None
    po: Optional[str] = None
    remarks: Optional[str] = None
    raw_row: dict[str, Any] = field(default_factory=dict)


@dataclass
class ParsedSheet:
    column_map: dict[str, str]
    rows: list[ParsedRow]
    skipped_rows: int
    warnings: list[str] = field(default_factory=list)


def _norm(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).replace("_", " ")
    return re.sub(r"\s+", " ", s).strip().lower()


def _to_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s in ("", "-"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _detect_columns(header_cells: list[tuple[int, Any]]) -> dict[str, int]:
    """header_cells: list of (col_index, raw_value). Returns field -> col_index."""
    scored: dict[str, list[tuple[int, int]]] = {}
    for col, raw in header_cells:
        text = _norm(raw)
        if not text:
            continue
        for fname, aliases in ALIASES.items():
            best = 0
            for a in aliases:
                if text == a:
                    best = max(best, 100 + len(a))
                elif a in text:
                    best = max(best, len(a))
            if best:
                scored.setdefault(fname, []).append((best, col))
    cols: dict[str, int] = {}
    used: set[int] = set()
    # assign highest-scoring field/column pairs first
    flat = sorted(
        ((score, fname, col) for fname, lst in scored.items() for score, col in lst),
        reverse=True,
    )
    for _score, fname, col in flat:
        if fname in cols or col in used:
            continue
        cols[fname] = col
        used.add(col)
    return cols


def _value_fallback_code_col(ws, header_row: int, max_col: int, sample: int = 60) -> Optional[int]:
    """Pick the column whose values look most like item codes (for junk headers)."""
    counts: dict[int, int] = {}
    seen = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        seen += 1
        for c in range(max_col):
            v = row[c] if c < len(row) else None
            if v is not None and CODE_RE.match(str(v).strip()):
                counts[c + 1] = counts.get(c + 1, 0) + 1
        if seen >= sample:
            break
    if not counts:
        return None
    best_col = max(counts, key=lambda k: counts[k])
    if counts[best_col] >= max(3, seen * 0.3):
        return best_col
    return None


def _parse_worksheet(ws) -> tuple[list[ParsedRow], int, Optional[str]]:
    """Returns (rows, skipped, warning)."""
    max_col = ws.max_column or 0
    if max_col == 0:
        return [], 0, f"'{ws.title}': empty sheet."

    # Find the header row: the row (within first 8) that yields the most fields
    # AND includes a quantity column.
    best_row = 0
    best_cols: dict[str, int] = {}
    for r in range(1, min(ws.max_row, 8) + 1):
        header_cells = [(c, ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        cols = _detect_columns(header_cells)
        has_qty = "pending_qty" in cols or "total_qty" in cols
        if has_qty and len(cols) > len(best_cols):
            best_row, best_cols = r, cols

    if not best_row:
        return [], 0, f"'{ws.title}': no recognizable quantity column — skipped."

    cols = best_cols
    if "item_code" not in cols:
        fb = _value_fallback_code_col(ws, best_row, max_col)
        if fb:
            cols["item_code"] = fb
    if "item_code" not in cols:
        return [], 0, f"'{ws.title}': no item-code column found — skipped."

    # required qty source: pending preferred, else total.
    qty_field = "pending_qty" if "pending_qty" in cols else "total_qty"
    header_text = {
        f: str(ws.cell(row=best_row, column=cols[f]).value or "").strip() for f in cols
    }

    def cell(row: tuple, f: str):
        idx = cols.get(f)
        return row[idx - 1] if idx and idx - 1 < len(row) else None

    rows: list[ParsedRow] = []
    skipped = 0
    for i, row in enumerate(
        ws.iter_rows(min_row=best_row + 1, values_only=True), start=best_row + 1
    ):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        raw_code = cell(row, "item_code")
        code = str(raw_code).strip() if raw_code is not None else None
        qty = _to_number(cell(row, qty_field))

        # Skip rows with no code, or whose (pending) quantity is not positive.
        #
        # Pending = required - stock on hand, so zero means "exactly covered"
        # and negative means "surplus" (e.g. need 10,240, hold 26,832 →
        # -16,592). Neither is something to purchase, and a negative stored as
        # a requirement is actively harmful: it nets off genuine demand from
        # *other* items in any category total, and makes the reconciliation
        # maths (expected_max = ceil(required/moq)*moq) meaningless.
        #
        # The full original row is still preserved in `raw_row`, so the surplus
        # is not lost — it just isn't demand.
        if not code or qty is None or qty <= 0:
            skipped += 1
            continue

        raw_row = {header_text[f]: (row[cols[f] - 1] if cols[f] - 1 < len(row) else None) for f in cols}
        raw_row["_tab"] = ws.title

        def txt(f: str) -> Optional[str]:
            v = cell(row, f)
            return str(v).strip() if v is not None and str(v).strip() != "" else None

        rows.append(
            ParsedRow(
                row_number=i,
                raw_code=code,
                item_name=txt("item_name"),
                required_qty=qty,
                raw_unit=txt("unit"),
                lot=txt("lot"),
                department=txt("department"),
                color=txt("color"),
                location=ws.title,  # tab = plant / product line
                po=txt("po"),
                remarks=txt("remark"),
                raw_row=raw_row,
            )
        )
    return rows, skipped, None


def parse_rm_sheet(data: bytes) -> ParsedSheet:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)

    all_rows: list[ParsedRow] = []
    skipped_total = 0
    warnings: list[str] = []
    parsed_tabs: list[str] = []

    for ws in wb.worksheets:
        rows, skipped, warn = _parse_worksheet(ws)
        skipped_total += skipped
        if warn:
            warnings.append(warn)
        if rows:
            parsed_tabs.append(f"{ws.title} ({len(rows)})")
            all_rows.extend(rows)
    wb.close()

    if not all_rows:
        # Distinguish "we couldn't read this file" from "we read it fine and
        # there is simply nothing to purchase" — otherwise a fully-stocked
        # sheet reports a column-detection error, which sends the uploader
        # hunting for a formatting problem that doesn't exist.
        if skipped_total:
            raise ValueError(
                f"Nothing to purchase: all {skipped_total} row(s) have a "
                "pending quantity of zero or less, meaning stock on hand "
                "already covers the requirement."
            )
        raise ValueError(
            "No requirement rows found. Expected an item-code column (e.g. "
            "'COMPONENT ICODE') and a quantity column (e.g. 'Pending' / "
            "'TOTAL_COMPONENT_QUANTITY')."
        )

    warnings.insert(0, "Parsed tabs: " + ", ".join(parsed_tabs))
    return ParsedSheet(
        column_map={"tabs": ", ".join(parsed_tabs)},
        rows=all_rows,
        skipped_rows=skipped_total,
        warnings=warnings,
    )
