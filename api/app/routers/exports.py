"""Reconciliation export for the approval stage.

Produces an .xlsx the approver and MD can circulate: a KPI summary first, then
one tab per reconciliation status so a reviewer can drill into whichever KPI
looks wrong without filtering a single flat sheet.
"""
from __future__ import annotations

import datetime
import io
from typing import Any
from urllib.parse import quote

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse

from app.ratelimit import limit_exports
from app.auth import CurrentUser, get_current_user, require_roles
from app.supabase_client import fetch_all

router = APIRouter(prefix="/exports", tags=["exports"])

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

# Order matters: worst news first, so a reviewer meets the problems before the
# lines that are already fine.
STATUS_ORDER: list[tuple[str, str]] = [
    ("over_buy", "Over-buy"),
    ("partial", "Partial"),
    ("not_bought", "Not bought"),
    ("extra_not_in_sheet", "Extra (not in sheet)"),
    ("on_target", "On target"),
]

DETAIL_COLUMNS = [
    ("item_code", "Item code"),
    ("name", "Item name"),
    ("category", "Category"),
    ("location", "Plant"),
    ("lot", "Lot"),
    ("base_unit", "Unit"),
    ("required", "Required"),
    ("ordered", "Ordered"),
    ("variance", "Variance"),
    ("variance_pct", "Variance %"),
    ("status", "Status"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F2937")
HEAD_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(bottom=THIN)


def _num(v: Any) -> float:
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


def _autosize(ws, widths: dict[int, int]) -> None:
    for idx, width in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = width


def _write_header(ws, headers: list[str], row: int = 1) -> None:
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.fill = HEAD_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


@router.get("/rm-sheets/{sheet_id}/reconciliation.xlsx")
def export_reconciliation(
    sheet_id: str,
    user: CurrentUser = Depends(get_current_user),
):
    """KPI-oriented reconciliation workbook for the approval stage."""
    require_roles(user, "approver", "md", "purchase_head", "po_team")
    limit_exports(user.id)

    sheet_res = (
        user.client.table("rm_sheet")
        .select("id, style_ref, status, created_at")
        .eq("id", sheet_id)
        .limit(1)
        .execute()
    )
    if not sheet_res.data:
        raise HTTPException(404, "RM sheet not found.")
    sheet = sheet_res.data[0]
    style_ref = sheet.get("style_ref") or sheet_id[:8]

    rows = fetch_all(
        lambda: user.client.table("reconciliation")
        .select("*")
        .eq("rm_sheet_id", sheet_id),
        order_by=("rm_sheet_id", "item_code", "lot", "location"),
    )
    if not rows:
        raise HTTPException(422, "Nothing to export — this sheet has no reconciled lines.")

    by_status: dict[str, list[dict[str, Any]]] = {k: [] for k, _ in STATUS_ORDER}
    for r in rows:
        by_status.setdefault(r.get("status") or "unknown", []).append(r)

    wb = openpyxl.Workbook()

    # ---------- Summary (KPIs) ----------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Reconciliation — {style_ref}"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Sheet status: {sheet.get('status')}"
    ws["A3"] = (
        "Generated "
        f"{datetime.datetime.now().strftime('%Y-%m-%d %H:%M')} by {user.email or user.id}"
    )
    for r in (2, 3):
        ws.cell(row=r, column=1).font = Font(size=9, color="6B7280")

    _write_header(
        ws,
        ["KPI", "Lines", "% of lines", "Required", "Ordered", "Variance"],
        row=5,
    )

    total_lines = len(rows)
    row_i = 6
    for key, label in STATUS_ORDER:
        group = by_status.get(key, [])
        req = sum(_num(g.get("required")) for g in group)
        order = sum(_num(g.get("ordered")) for g in group)
        ws.cell(row=row_i, column=1, value=label)
        ws.cell(row=row_i, column=2, value=len(group))
        pct = ws.cell(
            row=row_i,
            column=3,
            value=(len(group) / total_lines) if total_lines else 0,
        )
        pct.number_format = "0.0%"
        for col, val in ((4, req), (5, order), (6, order - req)):
            c = ws.cell(row=row_i, column=col, value=val)
            c.number_format = "#,##0.00"
        for col in range(1, 7):
            ws.cell(row=row_i, column=col).border = BORDER
        row_i += 1

    total_req = sum(_num(r.get("required")) for r in rows)
    total_ord = sum(_num(r.get("ordered")) for r in rows)
    ws.cell(row=row_i, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row_i, column=2, value=total_lines).font = Font(bold=True)
    for col, val in ((4, total_req), (5, total_ord), (6, total_ord - total_req)):
        c = ws.cell(row=row_i, column=col, value=val)
        c.number_format = "#,##0.00"
        c.font = Font(bold=True)

    # Headline KPIs — the two an approver is actually judged on.
    on_target = len(by_status.get("on_target", []))
    row_i += 2
    for label, value, fmt in (
        ("Fulfilment rate (on target / all lines)",
         (on_target / total_lines) if total_lines else 0, "0.0%"),
        ("Lines with no PO raised", len(by_status.get("not_bought", [])), "#,##0"),
    ):
        ws.cell(row=row_i, column=1, value=label).font = Font(bold=True)
        c = ws.cell(row=row_i, column=2, value=value)
        c.number_format = fmt
        c.font = Font(bold=True)
        row_i += 1

    _autosize(ws, {1: 40, 2: 12, 3: 12, 4: 16, 5: 16, 6: 16})

    # ---------- One tab per KPI ----------
    for key, label in STATUS_ORDER:
        group = by_status.get(key, [])
        tab = wb.create_sheet(label[:31])
        _write_header(tab, [h for _, h in DETAIL_COLUMNS])
        for i, r in enumerate(group, start=2):
            req = _num(r.get("required"))
            for c, (field, _) in enumerate(DETAIL_COLUMNS, start=1):
                if field == "variance_pct":
                    # No requirement means no denominator — an extra line has
                    # nothing to be a percentage of.
                    v = (_num(r.get("variance")) / req) if req else None
                else:
                    v = r.get(field)
                cell = tab.cell(row=i, column=c, value=v)
                if field in ("required", "ordered", "variance"):
                    cell.number_format = "#,##0.00"
                elif field == "variance_pct":
                    cell.number_format = "+0.0%;-0.0%;0.0%"
        if not group:
            tab.cell(row=2, column=1, value="— none —").font = Font(
                italic=True, color="6B7280"
            )
        _autosize(
            tab,
            {1: 14, 2: 34, 3: 20, 4: 18, 5: 14, 6: 10, 7: 14, 8: 14, 9: 14,
             10: 10, 11: 14, 12: 12, 13: 18},
        )

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    buf.seek(0)

    user.client.table("audit_log").insert(
        {
            "actor_id": user.id,
            "entity": "rm_sheet",
            "entity_id": sheet_id,
            "action": "reconciliation_exported",
            "detail": {"lines": total_lines, "style_ref": style_ref},
        }
    ).execute()

    filename = f"reconciliation-{style_ref}.xlsx".replace(" ", "-")
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={
            # RFC 5987 — style_ref may contain non-ASCII.
            "Content-Disposition": (
                f"attachment; filename=\"{filename}\"; "
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )


# ---------------------------------------------------------------------------
# Registers
#
# Both are flat single-sheet workbooks rather than the tabbed KPI report: these
# get filtered and pivoted by whoever receives them, and a tabbed workbook
# fights that.
# ---------------------------------------------------------------------------

# Matches the screens. Deliveries are rarely exact — cloth is cut to roll
# length, hardware ships in carton multiples — so only an excess beyond this
# counts as over-received.
OVER_TOLERANCE = 0.02
SHORT_TOLERANCE = 0.005


def _ordered_and_received(user: CurrentUser) -> tuple[dict, dict, dict]:
    """Ordered per (PO number, item), received likewise, and PO detail by number.

    Keyed without the lot: the register and the RM sheet name lots from
    different systems, so including it meant a receipt never matched its order.
    """
    pos = fetch_all(
        lambda: user.client.table("po")
        .select("id, po_number, etd, site, status, rm_sheet_id, "
                "po_line(item_code, lot, ordered_qty, rate, supplier)")
        .neq("status", "draft"),
        order_by="id",
    )

    ordered: dict[tuple[str, str], float] = {}
    po_by_number: dict[str, dict] = {}
    for p in pos:
        num = (p.get("po_number") or "").upper()
        if num:
            po_by_number[num] = p
        for line in p.get("po_line") or []:
            code = (line.get("item_code") or "").upper()
            if not num or not code:
                continue
            key = (num, code)
            ordered[key] = ordered.get(key, 0.0) + _num(line.get("ordered_qty"))

    received: dict[tuple[str, str], float] = {}
    grn = fetch_all(
        lambda: user.client.table("grn_ours").select("po_number, item_code, qty"),
        order_by="id",
    )
    for g in grn:
        num = (g.get("po_number") or "").upper()
        code = (g.get("item_code") or "").upper()
        if not num or not code:
            continue
        key = (num, code)
        received[key] = received.get(key, 0.0) + _num(g.get("qty"))

    return ordered, received, po_by_number


@router.get("/grn-register.xlsx")
def export_grn_register(user: CurrentUser = Depends(get_current_user)):
    """Every receipt, with what it was ordered against.

    A receipt on its own is not reviewable — 3,085 arrived is only meaningful
    beside the order — so the comparison travels with the row rather than being
    something the recipient has to reconstruct with a lookup.
    """
    require_roles(user, "approver", "md", "purchase_head", "po_team")
    limit_exports(user.id)

    rows = fetch_all(
        lambda: user.client.table("grn_ours").select(
            "grc_no, grc_date, po_number, po_date, item_code, item_name, lot, "
            "qty, supplier, stock_point, department, doc_no, doc_date, "
            "landed_cost, remarks"
        ),
        order_by="id",
    )
    if not rows:
        raise HTTPException(422, "Nothing to export — no receipts have been imported.")

    ordered, received, _ = _ordered_and_received(user)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GRN register"

    headers = [
        "GRC no", "GRC date", "PO number", "PO date", "Item code", "Item name",
        "Lot", "This receipt", "Ordered", "Received (all receipts)", "Excess",
        "Supplier", "Stock point", "Department", "Doc no", "Doc date",
        "Landed cost", "Remarks",
    ]
    _write_header(ws, headers)

    for i, r in enumerate(rows, start=2):
        num = (r.get("po_number") or "").upper()
        code = (r.get("item_code") or "").upper()
        key = (num, code)
        ordr = ordered.get(key)
        recd = received.get(key, _num(r.get("qty")))
        excess = (
            recd - ordr if ordr and ordr > 0 and recd > ordr * (1 + OVER_TOLERANCE) else None
        )

        values = [
            r.get("grc_no"), r.get("grc_date"), r.get("po_number"), r.get("po_date"),
            r.get("item_code"), r.get("item_name"), r.get("lot"),
            _num(r.get("qty")),
            # Blank rather than zero when nothing matched: a zero would read as
            # "ordered none", which is a different and wrong statement.
            ordr if ordr is not None else "no match",
            recd, excess, r.get("supplier"), r.get("stock_point"),
            r.get("department"), r.get("doc_no"), r.get("doc_date"),
            r.get("landed_cost"), r.get("remarks"),
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            if c in (8, 10, 11) or (c == 9 and isinstance(v, (int, float))):
                cell.number_format = "#,##0.00"

    _autosize(ws, {1: 24, 3: 24, 5: 14, 6: 34, 7: 18, 12: 30, 13: 34})
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.date.today().isoformat()
    filename = f"grn-register-{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"; '
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )


@router.get("/approved-pos.xlsx")
def export_approved_pos(user: CurrentUser = Depends(get_current_user)):
    """Every purchase order the approver has signed off, with when.

    One row per PO, not per line: approval is recorded once per order, so the
    date and time travel with the order rather than being repeated a line at a
    time.
    """
    require_roles(user, "approver", "md", "purchase_head", "po_team")
    limit_exports(user.id)

    pos = fetch_all(
        lambda: user.client.table("po")
        .select(
            "po_number, site, etd, approved_at, approved_by, approval_note, "
            "rm_sheet_id, rm_sheet(style_ref), "
            "po_line(item_code, lot, ordered_qty, rate, supplier)"
        )
        .eq("approval_status", "approved"),
        order_by="id",
    )
    if not pos:
        raise HTTPException(
            422, "Nothing to export — no purchase order has been approved yet."
        )

    approver_ids = list({p["approved_by"] for p in pos if p.get("approved_by")})
    names: dict[str, str] = {}
    if approver_ids:
        profs = (
            user.client.table("profiles")
            .select("id, full_name, email")
            .in_("id", approver_ids)
            .execute()
            .data
        ) or []
        names = {
            pr["id"]: pr.get("full_name") or pr.get("email") or pr["id"][:8]
            for pr in profs
        }

    # Most recently approved first.
    pos.sort(key=lambda p: p.get("approved_at") or "", reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved POs"
    _write_header(ws, [
        "PO number", "Sheet", "Supplier", "Site", "ETD", "Ordered qty",
        "Value", "Approved by", "Approved date", "Approved time", "Note",
    ])

    for i, p in enumerate(pos, start=2):
        lines = p.get("po_line") or []
        supplier = ", ".join(
            sorted({s for l in lines if (s := (l.get("supplier") or "").strip())})
        ) or None
        ordered = sum(_num(l.get("ordered_qty")) for l in lines)
        value = sum(
            _num(l.get("ordered_qty")) * _num(l.get("rate"))
            for l in lines
            if l.get("rate") is not None
        ) or None

        sheet = p.get("rm_sheet")
        if isinstance(sheet, list):
            sheet = sheet[0] if sheet else None
        style_ref = (sheet or {}).get("style_ref")

        # Split once, here, rather than leaving date and time bundled in one
        # ISO string a reviewer would otherwise have to parse by eye.
        approved_at = p.get("approved_at")
        approved_date = approved_time = None
        if approved_at:
            try:
                dt = datetime.datetime.fromisoformat(str(approved_at).replace("Z", "+00:00"))
                approved_date = dt.date()
                approved_time = dt.strftime("%H:%M:%S")
            except ValueError:
                pass

        values = [
            p.get("po_number"), style_ref, supplier, p.get("site"), p.get("etd"),
            ordered, value,
            names.get(p.get("approved_by")) if p.get("approved_by") else None,
            approved_date, approved_time, p.get("approval_note"),
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            if c in (6, 7):
                cell.number_format = "#,##0.00"
            if c == 9 and v is not None:
                cell.number_format = "yyyy-mm-dd"

    _autosize(ws, {1: 20, 2: 16, 3: 30, 4: 20, 5: 14, 8: 24, 11: 30})
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    today = datetime.date.today().isoformat()
    filename = f"approved-pos-{today}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"; '
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )


@router.get("/pending-pos.xlsx")
def export_pending_pos(user: CurrentUser = Depends(get_current_user)):
    """Purchase orders with material still outstanding, line by line.

    One row per PO line rather than per PO: chasing a supplier is about a
    specific material, and a PO-level total does not say which item is short.
    """
    require_roles(user, "approver", "md", "purchase_head", "po_team")
    limit_exports(user.id)

    ordered, received, po_by_number = _ordered_and_received(user)

    sheets = {
        s["id"]: s.get("style_ref")
        for s in fetch_all(
            lambda: user.client.table("rm_sheet").select("id, style_ref"),
            order_by="id",
        )
    }

    today = datetime.date.today()
    out: list[list[Any]] = []

    for num, p in po_by_number.items():
        etd_raw = p.get("etd")
        etd = None
        days_late = None
        if etd_raw:
            try:
                etd = datetime.date.fromisoformat(str(etd_raw))
                days_late = max(0, (today - etd).days)
            except ValueError:
                etd = None

        for line in p.get("po_line") or []:
            code = (line.get("item_code") or "").upper()
            if not code:
                continue
            ordr = _num(line.get("ordered_qty"))
            if ordr <= 0:
                continue
            recd = received.get((num, code), 0.0)
            # Complete within tolerance is not outstanding, whatever the
            # rounding in the register says.
            if recd >= ordr * (1 - SHORT_TOLERANCE):
                continue

            out.append([
                num,
                line.get("supplier"),
                p.get("site"),
                etd,
                days_late if (days_late or 0) > 0 else None,
                "overdue" if (days_late or 0) > 0 else ("not due yet" if etd else "no date"),
                sheets.get(p.get("rm_sheet_id")),
                line.get("item_code"),
                line.get("lot"),
                ordr,
                recd,
                ordr - recd,
                "nothing received" if recd == 0 else "part received",
            ])

    if not out:
        raise HTTPException(
            422, "Nothing to export — every purchase order has been fully received."
        )

    # Most overdue first, then the nearest date: the order they need chasing in.
    out.sort(key=lambda r: (-(r[4] or 0), str(r[3] or "9999-12-31")))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pending POs"
    _write_header(ws, [
        "PO number", "Supplier", "Delivery site", "ETD", "Days late", "Status",
        "Sheet", "Item code", "Lot", "Ordered", "Received", "Outstanding",
        "Group",
    ])

    for i, row in enumerate(out, start=2):
        for c, v in enumerate(row, start=1):
            cell = ws.cell(row=i, column=c, value=v)
            if c in (10, 11, 12):
                cell.number_format = "#,##0.00"
            if c == 4 and v is not None:
                cell.number_format = "yyyy-mm-dd"

    _autosize(ws, {1: 26, 2: 30, 3: 34, 6: 14, 7: 16, 8: 14, 9: 20, 13: 18})
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"pending-pos-{today.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{filename}"; '
                f"filename*=UTF-8''{quote(filename)}"
            )
        },
    )
