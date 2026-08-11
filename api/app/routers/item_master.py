import io
from typing import Any, Optional

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile

from app.uploads import read_upload
from app.ratelimit import limit_uploads
from app.auth import CurrentUser, get_current_user, require_roles
from app.parsing import _norm, _to_number

router = APIRouter(prefix="/item-master", tags=["item-master"])

# Exact (normalised) header -> field. Exact matching avoids "ITEM NAME" vs
# "ARTICLE NAME" colliding on the word "name".
HEADER_MAP: dict[str, str] = {
    # item_code
    "inv code": "item_code",
    "item code": "item_code",
    "itemcode": "item_code",
    "component icode": "item_code",
    "icode": "item_code",
    "sku": "item_code",
    "code": "item_code",
    # name
    "item name": "name",
    "component item name": "name",
    "name": "name",
    "description": "name",
    "particulars": "name",
    # article / category path
    "article name": "article_name",
    "article": "article_name",
    # explicit category (if present)
    "category": "category",
    "department": "category",
    "dept": "category",
    # unit
    "uom": "base_unit",
    "unit": "base_unit",
    "units": "base_unit",
    "base unit": "base_unit",
    # moq (usually absent — buyers set MOQ per PO line)
    "moq": "moq",
    "min order qty": "moq",
    "minimum order quantity": "moq",
    # reference-only
    "hsn code": "hsn_code",
    "hsn": "hsn_code",
    "hsn/sac": "hsn_code",
    "material type": "material_type",
    "material": "material_type",
    "type": "material_type",
}

# Normalise common unit spellings to a canonical base unit.
UOM_MAP: dict[str, str] = {
    "mtr": "meter", "mtrs": "meter", "mt": "meter", "m": "meter",
    "meter": "meter", "meters": "meter", "metre": "meter", "metres": "meter",
    "pc": "piece", "pcs": "piece", "pcs.": "piece", "piece": "piece",
    "pieces": "piece", "nos": "piece", "no": "piece", "pcs ": "piece",
    "kg": "kg", "kgs": "kg", "kilogram": "kg", "kilograms": "kg",
    "roll": "roll", "rolls": "roll",
    "set": "set", "sets": "set",
    "box": "box", "boxes": "box",
    "pair": "pair", "pairs": "pair",
    "cone": "cone", "cones": "cone",
    "carton": "carton", "ctn": "carton",
    "bag": "bag", "bags": "bag",
}


def _norm_unit(v: Any) -> str:
    n = _norm(v)
    return UOM_MAP.get(n, n) if n else "piece"


def _map_columns(ws) -> tuple[int, dict[str, int]]:
    for r in range(1, min(ws.max_row, 10) + 1):
        cols: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            field_ = HEADER_MAP.get(_norm(ws.cell(row=r, column=c).value))
            if field_ and field_ not in cols:
                cols[field_] = c
        if "item_code" in cols:
            return r, cols
    raise ValueError(
        "Could not find an item-code column (e.g. 'INV CODE' / 'Item Code')."
    )


@router.post("/import")
def import_item_master(
    file: UploadFile = File(...),
    user: CurrentUser = Depends(get_current_user),
):
    require_roles(user, "admin")

    if not (file.filename or "").lower().endswith(".xlsx"):
        raise HTTPException(400, "Please upload an .xlsx file.")
    limit_uploads(user.id)
    data = read_upload(file, "item master file")

    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.worksheets[0]
    try:
        header_row, cols = _map_columns(ws)
    except ValueError as e:
        wb.close()
        raise HTTPException(422, str(e))

    def get(row, field_) -> Optional[Any]:
        idx = cols.get(field_)
        return row[idx - 1] if idx and idx - 1 < len(row) else None

    def text(v: Any) -> Optional[str]:
        return str(v).strip() if v is not None and str(v).strip() != "" else None

    seen: dict[str, dict[str, Any]] = {}
    skipped = 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue
        code = text(get(row, "item_code"))
        if not code:
            skipped += 1
            continue

        article = text(get(row, "article_name"))
        explicit_cat = text(get(row, "category"))
        # Full-path category: the whole ARTICLE NAME, else an explicit category.
        category = explicit_cat or article or "uncategorized"
        moq = _to_number(get(row, "moq"))

        seen[code.upper()] = {
            "item_code": code,
            "name": text(get(row, "name")) or code,
            "category": category,
            "base_unit": _norm_unit(get(row, "base_unit")),
            "moq": moq if moq is not None else 0,
            "article_name": article,
            "hsn_code": text(get(row, "hsn_code")),
            "material_type": text(get(row, "material_type")),
        }
    wb.close()

    rows = list(seen.values())
    if not rows:
        raise HTTPException(422, "No item rows found.")

    CHUNK = 500
    for i in range(0, len(rows), CHUNK):
        user.client.table("item_master").upsert(
            rows[i : i + CHUNK], on_conflict="item_code"
        ).execute()

    user.client.table("audit_log").insert(
        {
            "actor_id": user.id,
            "entity": "item_master",
            "entity_id": None,
            "action": "imported",
            "detail": {"rows": len(rows), "skipped": skipped, "columns": cols},
        }
    ).execute()

    return {
        "imported": len(rows),
        "skipped": skipped,
        "detected_columns": list(cols.keys()),
    }
