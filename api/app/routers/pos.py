import hashlib
import io
import re
from datetime import datetime, timezone
from typing import Any, Optional
import openpyxl

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile

from app.auth import CurrentUser, get_current_user, require_roles
from app.po_number import po_number_from_pdf
from app.routers.notify import notify_po_attached
from app.supabase_client import fetch_all

router = APIRouter(prefix="/pos", tags=["pos"])


def _explain_po_insert(e: Exception) -> str:
    """Why creating a PO row failed, in words the importer can act on."""
    t = str(e)
    if "23505" in t or "po_po_number_key" in t or "duplicate key" in t.lower():
        return "a purchase order with that number already exists"
    return t[:200]

ALLOWED_EXT = (".pdf", ".xlsx", ".xls", ".png", ".jpg", ".jpeg")

PO_HEADER_ALIASES: dict[str, list[str]] = {
    "po_number": ["po number", "po no", "po #", "ponumber", "po", "po_id"],
    "item_code": ["component icode", "icode", "item code", "itemcode", "sku", "code", "item"],
    "ordered_qty": ["ordered qty", "order qty", "po qty", "qty", "quantity", "ordered"],
    "moq": ["moq", "minimum order quantity", "moq qty"],
    "supplier": ["supplier", "vendor", "party name", "party", "supplier name"],
    # The reconciliation view joins requirements to PO lines on
    # (item_code, lot, location), so these must survive the import.
    "lot": ["lot no", "lot number", "lot"],
    "location": ["location", "plant", "warehouse"],
}


def _norm(v: Any) -> str:
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v)).strip().lower()


def _to_number(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "")
    if s == "" or s == "-":
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _match_po_header(cell: str) -> Optional[str]:
    n = _norm(cell)
    if not n:
        return None
    for fieldname, aliases in PO_HEADER_ALIASES.items():
        if n in aliases:
            return fieldname
    for fieldname, aliases in sorted(
        PO_HEADER_ALIASES.items(),
        key=lambda kv: max(len(a) for a in kv[1]),
        reverse=True,
    ):
        for a in aliases:
            if len(a) >= 3 and a in n:
                return fieldname
    return None


@router.post("/{po_id}/upload")
def upload_po_document(
    po_id: str,
    file: UploadFile = File(...),
    po_number: str = Form(default=""),
    user: CurrentUser = Depends(get_current_user),
):
    """PO team attaches the finalised PO document to a PO and marks it uploaded.

    Also records the supplier-facing PO number. A number typed by the PO team
    always wins; otherwise it is read out of the PDF. This is the identifier the
    GRN register will refer to, so it is captured at the one moment the document
    is in front of the person who can confirm it.
    """
    require_roles(user, "po_team")

    filename = file.filename or "po"
    if not filename.lower().endswith(ALLOWED_EXT):
        raise HTTPException(400, f"Allowed types: {', '.join(ALLOWED_EXT)}")

    data = file.file.read()
    if not data:
        raise HTTPException(400, "Uploaded file is empty.")

    # Confirm the PO exists and is visible to this user (RLS: po_select).
    po = (
        user.client.table("po")
        .select("id, rm_sheet_id, doc_path, po_number")
        .eq("id", po_id)
        .limit(1)
        .execute()
    )
    if not po.data:
        raise HTTPException(404, "PO not found.")

    previous_path = po.data[0].get("doc_path")

    ext = "." + filename.lower().rsplit(".", 1)[-1]
    # Each upload gets its own object rather than overwriting `document.ext`.
    #
    # Attaching a second document used to replace the first in place — the
    # bytes of the signed PO an approver had already read were gone, with no
    # copy anywhere. Worse, replacing a .pdf with a .xlsx wrote to a different
    # name, so the original stayed in storage unreferenced while the record
    # pointed elsewhere: sometimes overwritten, sometimes orphaned, depending
    # on the file extension.
    #
    # The timestamp makes the path unique, so nothing is ever destroyed and the
    # audit trail's paths still resolve to the bytes they described.
    stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    path = f"{po_id}/{stamp}-document{ext}"

    content_type = {
        ".pdf": "application/pdf",
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xls": "application/vnd.ms-excel",
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
    }.get(ext, "application/octet-stream")

    try:
        # upsert deliberately off: the path is unique per upload, so a clash
        # means something is wrong and should be an error rather than a
        # silently discarded document.
        user.client.storage.from_("po-docs").upload(
            path, data, {"content-type": content_type, "upsert": "false"}
        )
    except Exception as e:
        raise HTTPException(502, f"Storage upload failed: {e}")

    # The typed number wins; otherwise read it off the document. Extraction is
    # best-effort and never blocks the attachment — a missing number is fixed
    # by typing it, a wrong one would quietly misdirect GRN matching.
    typed = po_number.strip().upper()
    extracted = po_number_from_pdf(data) if ext == ".pdf" else None
    resolved = typed or extracted

    fields: dict[str, Any] = {
        "doc_path": path,
        "uploaded_by": user.id,
        "status": "uploaded",
    }
    if resolved:
        fields["po_number"] = resolved

    # Replacing an attached document is allowed — a wrong or unsigned file gets
    # attached and has to be correctable — but it is said out loud. The
    # approver may already have read the document being replaced.
    warning: Optional[str] = (
        "This PO already had a document attached; it has been replaced. The "
        "previous file is kept and the change is recorded in the audit log."
        if previous_path
        else None
    )
    try:
        upd = user.client.table("po").update(fields).eq("id", po_id).execute()
    except Exception as e:
        # A PO number already claimed by another order must not cost us the
        # attachment. The file is in storage by this point, so failing the whole
        # update leaves the document uploaded but the PO showing none, and every
        # retry fails the same way. Attach it, and report the number separately
        # for someone to correct — which the PO number field on the card allows.
        if resolved and _is_duplicate_po_number(e):
            dupe = (
                f"The document is attached, but PO number {resolved} already "
                "belongs to another purchase order, so it was not saved. Check "
                "the number on the document and set it on this PO."
            )
            warning = f"{warning} {dupe}" if warning else dupe
            fields.pop("po_number", None)
            resolved = None
            try:
                upd = user.client.table("po").update(fields).eq("id", po_id).execute()
            except Exception as e2:
                raise HTTPException(500, f"Failed to update PO record: {e2}")
        else:
            raise HTTPException(500, f"Failed to update PO record: {e}")
    if not upd.data:
        raise HTTPException(500, "Failed to update PO record.")

    user.client.table("audit_log").insert(
        {
            "actor_id": user.id,
            "entity": "po",
            "entity_id": po_id,
            "action": "doc_uploaded",
            "detail": {
                "filename": filename,
                "path": path,
                # What it replaced, so the trail says which document an
                # approval was given against.
                "replaced_path": previous_path,
                "po_number": resolved,
                "po_number_source": "typed" if typed else ("extracted" if extracted else None),
            },
        }
    ).execute()

    # Hand off to the approver. Never fatal: the document is already stored.
    # A skipped send is folded into the warning the client already shows —
    # returning it in a field nothing reads is the same as saying nothing.
    notified: dict[str, Any]
    try:
        notified = notify_po_attached(user.client, po_id)
    except Exception as e:
        notified = {"sent": False, "error": str(e)}
    if not notified.get("sent"):
        why = notified.get("reason") or notified.get("detail") or notified.get("error")
        note = f"The approver was not notified: {why}" if why else (
            "The approver was not notified."
        )
        warning = f"{warning} {note}" if warning else note

    return {
        "po_id": po_id,
        "doc_path": path,
        "status": "uploaded",
        "po_number": resolved,
        "po_number_source": "typed" if typed else ("extracted" if extracted else None),
        "warning": warning,
        "notified": notified,
    }


def _is_duplicate_po_number(err: Exception) -> bool:
    """True when a write failed because another PO already holds the number.

    Matched on both the index name and the SQLSTATE, because the driver
    surfaces the error differently depending on whether PostgREST reports it as
    JSON or the client raises it directly.
    """
    text = str(err)
    return "po_po_number_key" in text or "23505" in text


@router.post("/import-register")
def import_po_register(
    sheet_id: str = Form(...),
    file: UploadFile = File(...),
    user: CurrentUser = Depends(get_current_user),
):
    """
    Automated PO Register Importer: Upload a PO Register (.xlsx) spreadsheet to
    automatically generate POs and PO line items without manual entry.
    """
    require_roles(user, "buyer", "po_team", "purchase_head")

    filename = file.filename or "po_register.xlsx"
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(400, "Please upload a .xlsx PO Register file.")

    data = file.file.read()
    if not data:
        raise HTTPException(400, "Uploaded file is empty.")

    # Idempotency: re-importing the same register would double every ordered
    # quantity and silently corrupt reconciliation. `po` has no content_hash
    # column, so the previous import is detected via its audit_log entry.
    content_hash = hashlib.sha256(data).hexdigest()
    prior = (
        user.client.table("audit_log")
        .select("id, detail")
        .eq("entity", "po")
        .eq("action", "po_register_imported")
        .eq("entity_id", sheet_id)
        .execute()
    )
    for row in prior.data or []:
        if (row.get("detail") or {}).get("content_hash") == content_hash:
            return {
                "sheet_id": sheet_id,
                "idempotent": True,
                "pos_created": 0,
                "lines_inserted": 0,
                "message": "This exact PO register was already imported for this sheet.",
            }

    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.worksheets[0]
    except Exception as e:
        raise HTTPException(400, f"Failed to read Excel workbook: {e}")

    # Locate header row
    header_row = 0
    cols: dict[str, int] = {}
    for r in range(1, min(ws.max_row, 15) + 1):
        found: dict[str, int] = {}
        for c in range(1, ws.max_column + 1):
            cell_val = ws.cell(row=r, column=c).value
            field_m = _match_po_header(str(cell_val)) if cell_val else None
            if field_m and field_m not in found:
                found[field_m] = c
        if "item_code" in found and "ordered_qty" in found:
            header_row = r
            cols = found
            break

    if not header_row:
        wb.close()
        raise HTTPException(
            422,
            "Could not locate required headers. Excel sheet must have an Item Code column (e.g., 'Item Code' or 'ICODE') and a Quantity column (e.g., 'Qty' or 'Ordered Qty').",
        )

    # Read data rows and group by PO Number
    po_groups: dict[str, list[dict[str, Any]]] = {}
    total_lines = 0

    for r_idx, row in enumerate(
        ws.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        if row is None or all(v is None or str(v).strip() == "" for v in row):
            continue

        item_code_val = row[cols["item_code"] - 1] if cols.get("item_code") and cols["item_code"] - 1 < len(row) else None
        qty_val = _to_number(row[cols["ordered_qty"] - 1]) if cols.get("ordered_qty") and cols["ordered_qty"] - 1 < len(row) else None

        if not item_code_val or qty_val is None or qty_val <= 0:
            continue

        item_code = str(item_code_val).strip()
        po_num = (
            str(row[cols["po_number"] - 1]).strip()
            if cols.get("po_number") and cols["po_number"] - 1 < len(row) and row[cols["po_number"] - 1]
            else f"PO-{sheet_id[:6]}-IMP"
        )
        moq = (
            _to_number(row[cols["moq"] - 1])
            if cols.get("moq") and cols["moq"] - 1 < len(row)
            else 0.0
        ) or 0.0

        def _text(field: str) -> Optional[str]:
            idx = cols.get(field)
            if not idx or idx - 1 >= len(row):
                return None
            v = row[idx - 1]
            return str(v).strip() if v is not None and str(v).strip() != "" else None

        if po_num not in po_groups:
            po_groups[po_num] = []

        po_groups[po_num].append({
            "item_code": item_code,
            "lot": _text("lot"),
            "location": _text("location"),
            "ordered_qty": qty_val,
            "moq": moq,
        })
        total_lines += 1

    wb.close()

    if not po_groups:
        raise HTTPException(422, "No valid PO lines found in the uploaded register.")

    # `po_line.item_code` has a foreign key to `item_master`; an unknown code
    # aborts the whole batch insert. Drop those lines and report them instead.
    catalogue = fetch_all(
        lambda: user.client.table("item_master").select("item_code"),
        order_by="item_code",
    )
    known = {
        (c["item_code"] or "").strip().upper(): c["item_code"] for c in catalogue
    }
    unknown_codes: set[str] = set()
    for po_num in list(po_groups):
        kept = []
        for line in po_groups[po_num]:
            canonical = known.get(line["item_code"].strip().upper())
            if canonical is None:
                unknown_codes.add(line["item_code"])
                continue
            line["item_code"] = canonical  # normalise to catalogue casing
            kept.append(line)
        if kept:
            po_groups[po_num] = kept
        else:
            del po_groups[po_num]

    if not po_groups:
        raise HTTPException(
            422,
            "None of the item codes in this register exist in the catalogue "
            f"(e.g. {', '.join(sorted(unknown_codes)[:5])}). Import the item "
            "master first.",
        )

    # If the register carries no lot/location columns, inherit them from the
    # sheet's requirement lines where the item resolves unambiguously —
    # otherwise the reconciliation join produces phantom not_bought / extra pairs.
    if not cols.get("lot") and not cols.get("location"):
        reqs = fetch_all(
            lambda: user.client.table("rm_requirement")
            .select("item_code, lot, location")
            .eq("rm_sheet_id", sheet_id)
            .not_.is_("item_code", "null"),
            order_by="id",
        )
        by_item: dict[str, list[dict[str, Any]]] = {}
        for r in reqs:
            by_item.setdefault((r["item_code"] or "").strip().upper(), []).append(r)
        for lines in po_groups.values():
            for line in lines:
                candidates = by_item.get(line["item_code"].strip().upper(), [])
                if len(candidates) == 1:
                    line["lot"] = candidates[0].get("lot")
                    line["location"] = candidates[0].get("location")

    # Create PO records and PO lines in database
    created_pos = 0
    inserted_lines = 0

    failed_pos: list[str] = []

    for po_num, lines in po_groups.items():
        # Wrapped, because po_number is unique. Re-importing a register that
        # overlaps one already loaded now raises instead of quietly creating a
        # second PO with the same number — which is the right refusal, but it
        # must cost that one order rather than the whole import.
        try:
            po_res = (
                user.client.table("po")
                .insert({
                    "rm_sheet_id": sheet_id,
                    "created_by": user.id,
                    "status": "uploaded",
                    # The number this PO is known by, saved on the row and not only
                    # in the line remark. Everything downstream joins on it: the GRN
                    # register, pending deliveries, the exports, and grn_ours — which
                    # hides receipts whose PO number matches nothing here. Without
                    # it these orders could never be matched to a delivery.
                    "po_number": po_num.strip().upper() or None,
                })
                .execute()
            )
        except Exception as e:
            failed_pos.append(f"{po_num} ({_explain_po_insert(e)})")
            continue
        if not po_res.data:
            failed_pos.append(po_num)
            continue

        po_id = po_res.data[0]["id"]
        lines_payload = [
            {
                "po_id": po_id,
                "item_code": l["item_code"],
                "lot": l.get("lot"),
                "location": l.get("location"),
                "ordered_qty": l["ordered_qty"],
                "moq": l["moq"],
                "remark": f"Auto-imported from PO Register ({po_num})",
            }
            for l in lines
        ]
        try:
            line_res = user.client.table("po_line").insert(lines_payload).execute()
        except Exception:
            # Roll back the empty PO — otherwise it reads as a real order with
            # no lines and skews the reconciliation counts.
            user.client.table("po").delete().eq("id", po_id).execute()
            failed_pos.append(po_num)
            continue

        created_pos += 1
        inserted_lines += len(line_res.data or [])

    # Audit log entry
    user.client.table("audit_log").insert({
        "actor_id": user.id,
        "entity": "po",
        "entity_id": sheet_id,
        "action": "po_register_imported",
        "detail": {
            "filename": filename,
            "content_hash": content_hash,
            "pos_created": created_pos,
            "lines_inserted": inserted_lines,
            "skipped_unknown_items": sorted(unknown_codes),
            "failed_pos": failed_pos,
        },
    }).execute()

    message = (
        f"Imported {inserted_lines} PO line(s) across {created_pos} Purchase Order(s)."
    )
    if unknown_codes:
        message += (
            f" Skipped {len(unknown_codes)} item code(s) not in the catalogue: "
            f"{', '.join(sorted(unknown_codes)[:5])}"
            f"{'...' if len(unknown_codes) > 5 else ''}."
        )
    if failed_pos:
        message += f" Failed to import: {', '.join(failed_pos[:5])}."

    return {
        "sheet_id": sheet_id,
        "idempotent": False,
        "pos_created": created_pos,
        "lines_inserted": inserted_lines,
        "skipped_unknown_items": sorted(unknown_codes),
        "failed_pos": failed_pos,
        "message": message,
    }
